import openpyxl
import sys
import converter.scheds as scheds
import os


def get_fam_name(cell_value):
    return cell_value.split('.')[0].strip()


def mimutes_to_str(minutes):
    return str(minutes // 60) + ':' + str(minutes % 60)


def isVil(text):
    if text is None:
        return False
    tmp = text.upper()
    for village in ('(ТЕРМАЛЬНЫЙ)', '(КОРЯКИ)', '(НИКОЛАЕВКА)', '(РАЗДОЛЬНЫЙ)', '(ПИОНЕРСКИЙ)', '(НАГОРНЫЙ)'):
        if tmp.find(village) > -1:
            return True
    else:
        return False


def get_sheetNames(file_name):
    swb = openpyxl.load_workbook(file_name)
    sheet_names = swb.get_sheet_names()
    sheet_names.sort()
    swb.close()
    return sheet_names


def convert(source_file_name, source_sheet_name, destination_file_name=r'C:\data\График за Июнь 2019.xlsx',
            destination_sheet_name='Лист1'):
    swb = openpyxl.load_workbook(filename=source_file_name)  # исходная книга
    source_sheet = swb[source_sheet_name]  # исходный лист

    dwb_filename = destination_file_name
    dwb = openpyxl.load_workbook(filename=dwb_filename)  # целевая книга
    destination_sheet = dwb[destination_sheet_name]  # целевой лист

    days = 0
    col = 3
    while type(source_sheet.cell(row=4, column=col).value) == int:
        days += 1
        col += 1

    fam_row_map = dict()
    cur_row = 5
    step = 3

    while destination_sheet.cell(row=cur_row, column=1).value is not None:
        tmp = get_fam_name(destination_sheet.cell(row=cur_row, column=2).value).split()
        fam_name = tmp[0] + ' ' + tmp[1][0]
        fam_row_map[fam_name] = cur_row
        cur_row += step

    schedule = dict()
    cur_row = 6
    step = 2

    while type(source_sheet.cell(row=cur_row, column=1).value) == int:
        text = source_sheet.cell(row=cur_row, column=2).value
        vilName = str(source_sheet.cell(row=cur_row, column=2).value)

        fam_name = get_fam_name(text)
        schedule[fam_name] = list()
        for day_number in range(1, days + 1):
            day = scheds.Day(source_sheet.cell(row=cur_row, column=day_number + 2).value,
                             str(source_sheet.cell(row=cur_row + 1, column=day_number + 2).value),
                             isVillage=isVil(vilName))
            # day.convert_from_eliz(source_sheet.cell(row=cur_row, column=day_number + 2).value, source_sheet.cell(row=cur_row + 1, column=day_number + 2).value)
            schedule[fam_name].append(day)
        cur_row += step

    destination_total_column_number = 1
    while True:
        if str(destination_sheet.cell(row=3, column=destination_total_column_number).value).find('Итого') > -1:
            break
        destination_total_column_number += 1

    for fam_name in fam_row_map.keys():
        destination_row = fam_row_map[fam_name]
        personal_schedule = schedule.get(fam_name)

        if personal_schedule is None:
            continue
        if personal_schedule is None:
            continue

        name = destination_sheet.cell(row=destination_row, column=2).value
        dayLen = 432 if 'аяь'.find(
            name[-1]) > -1 else 480  # Длительность производственного дня для Ж = 432, для М = 480

        total_work_days = 0
        total_time_in_minutes = 0
        day_number = 1
        for day in personal_schedule:
            if day.code == 'Я':
                destination_sheet.cell(row=destination_row, column=day_number + 3).value = day.start
                destination_sheet.cell(row=destination_row + 1, column=day_number + 3).value = day.fin
                destination_sheet.cell(row=destination_row + 2, column=day_number + 3).value = day.len
                total_work_days += 1
                try:
                    delta = day.get_len_in_minutes()
                    if delta == 0:
                        destination_sheet.cell(row=destination_row, column=day_number + 3).value = ''
                        destination_sheet.cell(row=destination_row + 1, column=day_number + 3).value = ''
                        destination_sheet.cell(row=destination_row + 2, column=day_number + 3).value = ''
                        total_work_days -= 1
                    total_time_in_minutes += delta
                except:
                    day_number += 1
                    continue
            else:
                destination_sheet.cell(row=destination_row, column=day_number + 3).value = day.code
                destination_sheet.cell(row=destination_row + 1, column=day_number + 3).value = ''
                destination_sheet.cell(row=destination_row + 2, column=day_number + 3).value = ''
            day_number += 1

        destination_sheet.cell(row=destination_row, column=destination_total_column_number).value = round(
            total_time_in_minutes / dayLen)
        destination_sheet.cell(row=destination_row + 2, column=destination_total_column_number).value = mimutes_to_str(
            total_time_in_minutes)

    newfile = dwb_filename[:-5] + '_NEW.xlsx'
    dwb.save(newfile)
    dwb.close()
    swb.close()
    os.startfile(newfile)
    sys.exit(0)
